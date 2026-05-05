"""
ProcDNA Intelligence — Export Engine v2
=========================================
Excel output — 4 sheets:
  1. Scenario Summary   — all pruned scenarios, 7 metrics, rank, confidence
  2. Territory Goals    — one row per (scenario × territory), final goals plain numbers
  3. Comparison         — best vs preferred side by side with delta
  4. Decision Report    — why chosen, trade-offs vs #2/#3, risk scenarios (rule-based)

CSV output — one row per (scenario × territory), flat.

Guardrails check — returns pass/fail per rule with explanation.
Delta insights   — attainment shift, goal difference, fairness change, rule-based narrative.
"""

import io
import csv
import numpy as np
import pandas as pd
from typing import List, Dict, Any, Optional

ATTAINMENT_LO = 0.85
ATTAINMENT_HI = 1.15


# ── Delta insights (Section X) ────────────────────────────────────────────────
def delta_insights(
    best: Dict,
    compare: Dict,
    national_forecast: float,
) -> Dict[str, Any]:
    """
    Structured delta between best scenario and a comparison scenario.
    Returns metric differences + 1-sentence rule-based narrative per key metric.
    """
    bf = best["fairness"]
    cf = compare["fairness"]

    att_shift   = round((cf["att"]  - bf["att"])  * 100, 1)
    fair_shift  = round(cf["score"] - bf["score"], 4)
    total_shift = round(cf["total_goals"] - bf["total_goals"], 2)
    cons_shift  = round((cf["cons"] - bf["cons"]) * 100, 1)
    ga_shift    = round((cf["Ga"]   - bf["Ga"])   * 100, 1)

    narratives = []

    if abs(att_shift) >= 1:
        direction = "improves" if att_shift > 0 else "reduces"
        narratives.append(
            f"Attainability {direction} by {abs(att_shift):.1f}pp — "
            f"{'more' if att_shift > 0 else 'fewer'} territories fall within the ±15% attainment band."
        )

    if abs(cons_shift) >= 2:
        direction = "more consistent" if cons_shift > 0 else "less consistent"
        narratives.append(
            f"Goals are {direction} across regions ({abs(cons_shift):.1f}pp shift in CV score)."
        )

    if abs(ga_shift) >= 2:
        direction = "better aligned to" if ga_shift > 0 else "less aligned to"
        narratives.append(
            f"Goals are {direction} territory growth trends ({abs(ga_shift):.1f}pp shift)."
        )

    if abs(total_shift) > 0:
        direction = "higher" if total_shift > 0 else "lower"
        pct = abs(total_shift / national_forecast * 100) if national_forecast > 0 else 0
        narratives.append(
            f"Total goals are {direction} by {abs(total_shift):,.0f} "
            f"({pct:.1f}% of national forecast)."
        )

    if abs(fair_shift) >= 0.01:
        direction = "better" if fair_shift > 0 else "lower"
        narratives.append(
            f"Overall fairness score is {direction} "
            f"({bf['score']:.3f} vs {cf['score']:.3f})."
        )

    return {
        "best_scenario_id":    best["scenario"]["id"],
        "compare_scenario_id": compare["scenario"]["id"],
        "attainability_shift_pp":    att_shift,
        "fairness_score_shift":      fair_shift,
        "total_goals_shift":         total_shift,
        "consistency_shift_pp":      cons_shift,
        "growth_alignment_shift_pp": ga_shift,
        "narrative": narratives if narratives else ["Scenarios are statistically similar across all metrics."],
    }


# ── Decision Report (Section XI) ─────────────────────────────────────────────
def _decision_report_rows(
    best: Dict,
    results: List[Dict],
    national_forecast: float,
) -> List[Dict]:
    """
    Build row data for Sheet 4 — Decision Report.
    Rule-based, no AI call.
    """
    bf    = best["fairness"]
    bsc   = best["scenario"]
    rows  = []

    # --- Section 1: Selected Plan ---
    rows.append({"Section": "SELECTED PLAN", "Item": "Scenario ID",       "Value": bsc["id"]})
    rows.append({"Section": "",               "Item": "Method",            "Value": bsc["method"]})
    rows.append({"Section": "",               "Item": "Plan Mode",         "Value": bsc["mode"]})
    rows.append({"Section": "",               "Item": "Preset Used",       "Value": bf.get("preset", "fairness")})
    rows.append({"Section": "",               "Item": "Composite Score",   "Value": bf["score"]})
    rows.append({"Section": "",               "Item": "Confidence",        "Value": bf.get("confidence", "—")})
    rows.append({"Section": "",               "Item": "On-Target %",       "Value": f"{bf['pct_on']*100:.1f}%"})
    rows.append({"Section": "",               "Item": "Total Goals",       "Value": bf["total_goals"]})
    rows.append({"Section": "",               "Item": "Forecast Gap",      "Value": bf["gap"]})

    # --- Section 2: Why it was chosen ---
    rows.append({"Section": "WHY CHOSEN", "Item": "", "Value": ""})
    metric_names = {"P": "Proportionality", "A": "Attainability", "C": "Consistency",
                    "Ga": "Growth Alignment", "V": "Volatility Penalty",
                    "K": "Cost Efficiency",  "S": "Goal Spread"}
    top3 = sorted(
        [(k, bf.get(k, 0)) for k in ["P","A","C","Ga","V","K","S"]],
        key=lambda x: -x[1]
    )[:3]
    for rank_i, (mk, mv) in enumerate(top3, 1):
        rows.append({
            "Section": "",
            "Item":    f"Top reason #{rank_i}: {metric_names.get(mk, mk)}",
            "Value":   f"{mv:.3f}",
        })

    # --- Section 3: Trade-offs vs #2 and #3 ---
    others = [r for r in results if not r["is_best"]][:2]
    rows.append({"Section": "TRADE-OFFS", "Item": "", "Value": ""})
    for alt in others:
        af  = alt["fairness"]
        aid = alt["scenario"]["id"]
        att_d = round((af["att"] - bf["att"]) * 100, 1)
        sc_d  = round(af["score"] - bf["score"], 4)
        tg_d  = round(af["total_goals"] - bf["total_goals"], 2)
        rows.append({
            "Section": f"vs {aid}",
            "Item":    "Attainability difference",
            "Value":   f"{'+' if att_d>=0 else ''}{att_d:.1f}pp",
        })
        rows.append({
            "Section": "",
            "Item":    "Fairness score difference",
            "Value":   f"{'+' if sc_d>=0 else ''}{sc_d:.4f}",
        })
        rows.append({
            "Section": "",
            "Item":    "Total goals difference",
            "Value":   f"{'+' if tg_d>=0 else ''}{tg_d:,.0f}",
        })

    # --- Section 4: Risk scenarios ---
    rows.append({"Section": "RISK SCENARIOS", "Item": "", "Value": ""})
    # Conservative: lowest total goals in top-10
    pool = results[:10]
    conservative = min(pool, key=lambda r: r["fairness"]["total_goals"])
    aggressive   = max(pool, key=lambda r: r["fairness"]["total_goals"])
    rows.append({
        "Section": "Conservative",
        "Item":    f"{conservative['scenario']['id']} — lowest total goals",
        "Value":   f"{conservative['fairness']['total_goals']:,.0f} "
                   f"(score {conservative['fairness']['score']:.3f})",
    })
    rows.append({
        "Section": "Aggressive",
        "Item":    f"{aggressive['scenario']['id']} — highest total goals",
        "Value":   f"{aggressive['fairness']['total_goals']:,.0f} "
                   f"(score {aggressive['fairness']['score']:.3f})",
    })

    return rows


# ── Guardrails (Section XII) ──────────────────────────────────────────────────
def guardrails_check(
    results: List[Dict],
    national_forecast: float,
    best_id: str,
) -> Dict[str, Any]:
    """
    Check the best scenario against 4 guardrail rules.
    Returns pass/fail per rule + explanation.
    """
    best = next((r for r in results if r["scenario"]["id"] == best_id), None)
    if not best:
        return {"error": "Best scenario not found"}

    goals = best["goals"]
    checks = []
    all_pass = True

    # Rule 1: Total goals within ±20% of national forecast
    if national_forecast > 0:
        total = sum(g["fg"] for g in goals)
        gap_pct = abs(total - national_forecast) / national_forecast * 100
        passed = gap_pct <= 20.0
        all_pass = all_pass and passed
        checks.append({
            "rule":        "Budget alignment",
            "description": "Total goals must be within ±20% of national forecast",
            "passed":      passed,
            "value":       f"{gap_pct:.1f}% deviation",
            "detail":      f"Total goals {total:,.0f} vs NF {national_forecast:,.0f}",
        })

    # Rule 2: No territory goal below 50% of average
    avg_goal = np.mean([g["fg"] for g in goals]) if goals else 1
    low_outliers = [g["tid"] for g in goals if g["fg"] < avg_goal * 0.5]
    passed = len(low_outliers) == 0
    all_pass = all_pass and passed
    checks.append({
        "rule":        "Minimum goal floor",
        "description": "No territory should have a goal below 50% of the average",
        "passed":      passed,
        "value":       f"{len(low_outliers)} territories below floor",
        "detail":      f"Outliers: {low_outliers[:5]}" if low_outliers else "None",
    })

    # Rule 3: No territory goal above 200% of average
    high_outliers = [g["tid"] for g in goals if g["fg"] > avg_goal * 2.0]
    passed = len(high_outliers) == 0
    all_pass = all_pass and passed
    checks.append({
        "rule":        "Maximum goal ceiling",
        "description": "No territory should have a goal above 200% of the average",
        "passed":      passed,
        "value":       f"{len(high_outliers)} territories above ceiling",
        "detail":      f"Outliers: {high_outliers[:5]}" if high_outliers else "None",
    })

    # Rule 4: Attainability — at least 70% of territories on target
    on_target_pct = sum(1 for g in goals if ATTAINMENT_LO <= g["att"] <= ATTAINMENT_HI) / max(len(goals), 1)
    passed = on_target_pct >= 0.70
    all_pass = all_pass and passed
    checks.append({
        "rule":        "Attainability floor",
        "description": "At least 70% of territories must have forecast within ±15% of goal",
        "passed":      passed,
        "value":       f"{on_target_pct*100:.1f}% on target",
        "detail":      f"{sum(1 for g in goals if ATTAINMENT_LO<=g['att']<=ATTAINMENT_HI)} / {len(goals)} territories",
    })

    return {
        "scenario_id": best_id,
        "all_pass":    all_pass,
        "checks":      checks,
        "summary":     "All guardrails passed." if all_pass else
                       f"{sum(1 for c in checks if not c['passed'])} guardrail(s) failed. Review before finalising.",
    }


# ── Excel export ──────────────────────────────────────────────────────────────
def to_excel(
    results: List[Dict],
    best_id: str,
    preferred_id: Optional[str],
    national_forecast: float,
    territories: list = None,
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    territories = territories or []
    terr_map = {t["territory_id"]: t for t in territories}

    HDR   = PatternFill("solid", start_color="1668B8", end_color="1668B8")
    DARK  = PatternFill("solid", start_color="0F1D2E", end_color="0F1D2E")
    GOLD  = PatternFill("solid", start_color="FFFDE7", end_color="FFFDE7")
    GREY  = PatternFill("solid", start_color="F1F4F8", end_color="F1F4F8")
    WHITE = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
    GRN   = PatternFill("solid", start_color="D4EDDA", end_color="D4EDDA")
    RED   = PatternFill("solid", start_color="F8D7DA", end_color="F8D7DA")
    AMB   = PatternFill("solid", start_color="FFF3CD", end_color="FFF3CD")
    BLU   = PatternFill("solid", start_color="E6F1FB", end_color="E6F1FB")
    thin  = Side(style="thin", color="D0D8E4")
    BDR   = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr(ws, r, c, v, w=12):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.fill = HDR; cell.border = BDR
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = w

    def cel(ws, r, c, v, fmt=None, bold=False, fill=None, align="left", color="000000"):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = Font(name="Arial", bold=bold, size=10, color=color)
        cell.border = BDR
        cell.alignment = Alignment(horizontal=align, vertical="center")
        if fill: cell.fill = fill
        if fmt:  cell.number_format = fmt
        return cell

    def title(ws, r, txt, ncols, sub=None):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        c = ws.cell(row=r, column=1, value=txt)
        c.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
        c.fill = DARK; c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = 26
        if sub:
            ws.merge_cells(start_row=r+1, start_column=1, end_row=r+1, end_column=ncols)
            s = ws.cell(row=r+1, column=1, value=sub)
            s.font = Font(name="Arial", italic=True, size=10, color="1668B8")
            s.fill = BLU; s.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[r+1].height = 18

    wb = Workbook()
    best_r = next((r for r in results if r["scenario"]["id"] == best_id), results[0] if results else None)

    # ── Sheet 1: EDA — Territory Baselines ───────────────────────────────────
    ws1 = wb.active; ws1.title = "1. EDA Summary"
    ws1.sheet_view.showGridLines = False; ws1.freeze_panes = "A5"
    n_terr = len(territories) or len(best_r["goals"]) if best_r else 0
    nat_h12 = sum(t.get("h12", 0) for t in territories)
    title(ws1, 1, "ProcDNA Intelligence — Territory Baseline Data", 14,
          f"{n_terr} territories | Quarterly NF: {national_forecast:,.0f} | National 12-month: {nat_h12:,.0f}")

    ws1.merge_cells("A3:N3")
    stats = (f"  Last Quarter National: {sum(t.get('q1',0) for t in territories):,.0f}  |  "
             f"H6: {sum(t.get('h6',0) for t in territories):,.0f}  |  "
             f"H9: {sum(t.get('h9',0) for t in territories):,.0f}  |  "
             f"H12: {nat_h12:,.0f}")
    c = ws1.cell(row=3, column=1, value=stats)
    c.font = Font(name="Arial", bold=True, size=10, color="1668B8")
    c.fill = BLU; c.alignment = Alignment(horizontal="left", vertical="center")
    ws1.row_dimensions[3].height = 20

    hs1 = [("Region ID",12),("Territory ID",14),("Territory Name",26),("Region Name",20),
           ("H6 Sales (6mo)",13),("H9 Sales (9mo)",13),("H12 Sales (12mo)",13),
           ("Q1 (Most Recent)",14),("Q2 (Prior)",12),("Q3",12),("Q4 (Oldest)",12),
           ("Market Share %",12),("Trend Score",11),("Volatility Band",15)]
    for i, (h, w) in enumerate(hs1, 1): hdr(ws1, 4, i, h, w)
    ws1.row_dimensions[4].height = 32

    for ri, t in enumerate(sorted(territories, key=lambda x: x.get("h12", 0), reverse=True), 5):
        fill = GREY if ri % 2 == 0 else WHITE
        ms   = t.get("h12", 0) / nat_h12 if nat_h12 > 0 else 0
        tr   = t.get("trend", 0)
        tcol = "1AA15A" if tr > 0.02 else ("D13B3B" if tr < -0.02 else "4A5B6E")
        vband = t.get("vband", "—")
        vfill = GRN if "Stable" in vband else (AMB if "Moderate" in vband else RED)
        cel(ws1, ri, 1,  t.get("region_id", ""),    fill=fill, align="center")
        cel(ws1, ri, 2,  t.get("territory_id", ""), fill=fill, align="center")
        cel(ws1, ri, 3,  t.get("territory_name", ""), fill=fill)
        cel(ws1, ri, 4,  t.get("region_name", ""),  fill=fill)
        cel(ws1, ri, 5,  t.get("h6", 0),  "#,##0", fill=fill, align="right")
        cel(ws1, ri, 6,  t.get("h9", 0),  "#,##0", fill=fill, align="right")
        cel(ws1, ri, 7,  t.get("h12", 0), "#,##0", fill=fill, align="right", bold=True)
        cel(ws1, ri, 8,  t.get("q1", 0),  "#,##0", fill=fill, align="right")
        cel(ws1, ri, 9,  t.get("q2", 0),  "#,##0", fill=fill, align="right")
        cel(ws1, ri, 10, t.get("q3", 0),  "#,##0", fill=fill, align="right")
        cel(ws1, ri, 11, t.get("q4", 0),  "#,##0", fill=fill, align="right")
        cel(ws1, ri, 12, ms,   "0.00%", fill=fill, align="center")
        cel(ws1, ri, 13, round(tr, 4), "0.0000", fill=fill, align="center", color=tcol, bold=tr > 0.02)
        cel(ws1, ri, 14, vband, fill=vfill, align="center")

    nr1 = len(territories) + 5
    ws1.merge_cells(start_row=nr1, start_column=1, end_row=nr1, end_column=4)
    c = ws1.cell(row=nr1, column=1, value="NATIONAL TOTAL")
    c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    c.fill = HDR; c.alignment = Alignment(horizontal="center", vertical="center"); c.border = BDR
    for ci in range(5, 12):
        cl = get_column_letter(ci)
        ws1.cell(row=nr1, column=ci, value=f"=SUM({cl}5:{cl}{nr1-1})")
        ws1.cell(row=nr1, column=ci).font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        ws1.cell(row=nr1, column=ci).fill = HDR
        ws1.cell(row=nr1, column=ci).number_format = "#,##0"
        ws1.cell(row=nr1, column=ci).border = BDR
        ws1.cell(row=nr1, column=ci).alignment = Alignment(horizontal="right")
    for col in [12, 13, 14]:
        ws1.cell(row=nr1, column=col).fill = HDR
        ws1.cell(row=nr1, column=col).border = BDR

    # ── Sheet 2: Forecasting ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("2. Forecasting")
    ws2.sheet_view.showGridLines = False; ws2.freeze_panes = "A4"
    title(ws2, 1, "ProcDNA Intelligence — Forecasting Results & Attainment", 9,
          "Best model selected per territory via wMAPE competition | Attainment = Forecast Q2 / Quarterly Goal")

    hs2 = [("Territory ID",14),("Territory Name",26),("Region",20),("Best Model",18),
           ("Forecast Q2 (Quarterly)",15),("wMAPE (Accuracy)",12),("Confidence",12),
           ("Quarterly Goal",14),("Attainment (Fc/Goal)",12)]
    for i, (h, w) in enumerate(hs2, 1): hdr(ws2, 3, i, h, w)
    ws2.row_dimensions[3].height = 32

    if best_r:
        goals_sorted = sorted(best_r["goals"], key=lambda g: g.get("fc", 0), reverse=True)
        for ri, g in enumerate(goals_sorted, 4):
            tid  = g["tid"]
            t    = terr_map.get(tid, {})
            att  = g.get("att", 0)
            on_t = 0.85 <= att <= 1.15
            fill = GREY if ri % 2 == 0 else WHITE
            att_fill = GRN if on_t else (AMB if att >= 0.5 else RED)
            wm   = t.get("wmape", 0.25)
            wcol = "1AA15A" if wm < 0.12 else ("D18A17" if wm < 0.20 else "D13B3B")
            cel(ws2, ri, 1, tid, fill=fill, align="center")
            cel(ws2, ri, 2, g.get("tname", ""), fill=fill)
            cel(ws2, ri, 3, g.get("rname", ""), fill=fill)
            cel(ws2, ri, 4, t.get("best_model", g.get("best_model","—")),
                fill=fill, bold=True, align="center", color="1668B8")
            cel(ws2, ri, 5, g.get("fc", 0), "#,##0", fill=fill, align="right", bold=True)
            cel(ws2, ri, 6, wm, "0.000", fill=fill, align="center", color=wcol, bold=wm < 0.12)
            cel(ws2, ri, 7, t.get("confidence", round(1-wm, 2)), "0.0%", fill=fill, align="center")
            cel(ws2, ri, 8, g["fg"], "#,##0", fill=fill, align="right")
            cel(ws2, ri, 9, att, "0.0%", fill=att_fill, align="center",
                bold=True, color="1AA15A" if on_t else ("D18A17" if att >= 0.5 else "D13B3B"))

    # ── Sheet 3: Scenario Scoring ─────────────────────────────────────────────
    ws3 = wb.create_sheet("3. Scenario Scoring")
    ws3.sheet_view.showGridLines = False; ws3.freeze_panes = "A6"
    title(ws3, 1, "ProcDNA Intelligence — Scenario Scoring (Top 25 After Pruning)", 20,
          f"NF = {national_forecast:,.0f} | P=Proportionality  A=Attainability  C=Consistency  Ga=Growth  V=Volatility  K=Cost  S=Spread")

    ws3.merge_cells("A3:T3")
    c = ws3.cell(row=3, column=1,
        value="  Composite = P×0.30 + A×0.25 + C×0.20 + Ga×0.10 + V×0.10 + K×0.03 + S×0.02  |  On-Target = Attainment ∈ [0.85, 1.15]  |  ★ = Best Scenario")
    c.font = Font(name="Arial", italic=True, size=9, color="4A5B6E")
    c.fill = GREY; c.alignment = Alignment(horizontal="left", vertical="center")
    ws3.row_dimensions[3].height = 16

    ws3.merge_cells("A4:T4")
    n_scored = len(results)
    on_best  = best_r["fairness"]["on_target"] if best_r else 0
    c = ws3.cell(row=4, column=1,
        value=f"  {n_scored} scenarios after Pareto pruning  |  Best scenario has {on_best}/{n_terr} territories on target  |  NF (quarterly): {national_forecast:,.0f}")
    c.font = Font(name="Arial", bold=True, size=9, color="1668B8")
    c.fill = BLU; c.alignment = Alignment(horizontal="left", vertical="center")
    ws3.row_dimensions[4].height = 18

    hs3 = [("★",5),("Rank",6),("Method",28),("Mode",12),
           ("W1 Q2wt",7),("W2 Q3wt",7),("W3 Q4wt",7),("Reg %",7),("Ind %",7),
           ("P Prop",8),("A Att",8),("C Con",8),("Ga Grw",8),
           ("V Vol",8),("K Cst",8),("S Sprd",8),
           ("Composite Score",13),("On Target",8),("On Target %",11),("Total Goals",13)]
    for i, (h, w) in enumerate(hs3, 1): hdr(ws3, 5, i, h, w)
    ws3.row_dimensions[5].height = 32

    for ri, r in enumerate(results, 6):
        sc  = r["scenario"]; fa = r["fairness"]
        is_best = (sc["id"] == best_id)
        fill = GOLD if is_best else (GREY if ri % 2 == 0 else WHITE)
        bold = is_best
        col  = "D18A17" if is_best else "000000"
        def pf(v): return round(float(v), 4) if v else 0

        cel(ws3, ri, 1,  "★" if is_best else "",   fill=fill, align="center", bold=True, color="D18A17")
        cel(ws3, ri, 2,  r["rank"],         "0",   fill=fill, align="center", bold=bold)
        cel(ws3, ri, 3,  sc["method"],             fill=fill, bold=bold)
        cel(ws3, ri, 4,  sc["mode"],               fill=fill, align="center")
        cel(ws3, ri, 5,  sc["w1"] or "—",  "0.0", fill=fill, align="center")
        cel(ws3, ri, 6,  sc["w2"] or "—",  "0.0", fill=fill, align="center")
        cel(ws3, ri, 7,  sc["w3"] or "—",  "0.0", fill=fill, align="center")
        cel(ws3, ri, 8,  round(sc.get("rs",0)*100),  "0", fill=fill, align="center")
        cel(ws3, ri, 9,  round(sc.get("is_",0)*100), "0", fill=fill, align="center")

        for ci, mk in enumerate(["P","A","C","Ga","V","K","S"], 10):
            v = pf(fa.get(mk, 0))
            if v >= 0.80:   mf = PatternFill("solid", start_color="C8E6C9", end_color="C8E6C9")
            elif v >= 0.60: mf = PatternFill("solid", start_color="FFF9C4", end_color="FFF9C4")
            elif v < 0.40:  mf = PatternFill("solid", start_color="FFCDD2", end_color="FFCDD2")
            else:           mf = fill
            cel(ws3, ri, ci, v, "0.000", fill=mf, align="center", bold=bold)

        cel(ws3, ri, 17, pf(fa["boosted_score"]), "0.0000",
            fill=fill, align="center", bold=True, color="1668B8" if is_best else "000000")
        cel(ws3, ri, 18, fa["on_target"],  "0",    fill=fill, align="center")
        cel(ws3, ri, 19, fa["pct_on"],     "0.0%", fill=fill, align="center")
        cel(ws3, ri, 20, fa["total_goals"],"#,##0",fill=fill, align="right")

    # ── Sheet 4: Best Scenario Goals ──────────────────────────────────────────
    ws4 = wb.create_sheet("4. Best Scenario — Goals")
    ws4.sheet_view.showGridLines = False; ws4.freeze_panes = "A8"
    if best_r:
        sc_b = best_r["scenario"]; fa_b = best_r["fairness"]
        wt_label = f"W1={sc_b['w1']} · W2={sc_b['w2']} · W3={sc_b['w3']}" if sc_b.get("w1") else "N/A"
        split_label = f"Regional {round(sc_b.get('rs',0)*100)}% / Individual {round(sc_b.get('is_',0)*100)}%"
        title(ws4, 1,
            f"THE ONE — {sc_b['method']} · {sc_b['mode']} · 2026Q2 Territory Goals", 14,
            f"Score {fa_b['boosted_score']:.4f} | {fa_b['on_target']}/{n_terr} on target "
            f"| {wt_label} | {split_label} | NF: {national_forecast:,.0f}")

        cards = [
            ("Method", sc_b["method"]), ("Mode", sc_b["mode"]),
            ("Weights", wt_label), ("Split", split_label),
            ("Composite", f"{fa_b['boosted_score']:.4f}"),
            ("On Target", f"{fa_b['on_target']} / {n_terr} ({fa_b['pct_on']*100:.1f}%)"),
            ("Total Goals", f"{fa_b['total_goals']:,.0f}"),
            ("NF (Quarterly)", f"{national_forecast:,.0f}"),
        ]
        for i, (k, v) in enumerate(cards):
            col = (i % 4) * 2 + 1; row = 3 + i // 4
            c1 = ws4.cell(row=row, column=col, value=k)
            c1.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
            c1.fill = HDR; c1.border = BDR
            c1.alignment = Alignment(horizontal="center", vertical="center")
            c2 = ws4.cell(row=row, column=col+1, value=v)
            c2.font = Font(name="Arial", bold=True, size=10, color="1668B8")
            c2.fill = BLU; c2.border = BDR
            c2.alignment = Alignment(horizontal="center", vertical="center")
        for row in [3, 4]: ws4.row_dimensions[row].height = 20
        ws4.row_dimensions[6].height = 8

        hs4 = [("Rank",6),("Territory ID",14),("Territory Name",26),("Region",20),
               ("Quarterly Goal",14),("Forecast Q2 2026",14),("Attainment (Fc/Goal)",13),
               ("On Target?",10),("Individual Goal",14),("Regional Goal",14),
               ("Best Model",18),("wMAPE",10),("Confidence",12),("Market Share %",12)]
        for i, (h, w) in enumerate(hs4, 1): hdr(ws4, 7, i, h, w)
        ws4.row_dimensions[7].height = 32

        for ri, g in enumerate(sorted(best_r["goals"], key=lambda x: x["fg"], reverse=True), 8):
            tid  = g["tid"]
            t    = terr_map.get(tid, {})
            att  = g.get("att", 0)
            on_t = 0.85 <= att <= 1.15
            fill = GREY if ri % 2 == 0 else WHITE
            att_fill = GRN if on_t else (AMB if att >= 0.5 else RED)
            ms = t.get("h12", 0) / nat_h12 if nat_h12 > 0 else 0
            wm = t.get("wmape", 0.25)
            cel(ws4, ri, 1,  ri-7,  "0", fill=fill, align="center", bold=True)
            cel(ws4, ri, 2,  tid,        fill=fill, align="center")
            cel(ws4, ri, 3,  g.get("tname",""), fill=fill)
            cel(ws4, ri, 4,  g.get("rname",""), fill=fill)
            cel(ws4, ri, 5,  g["fg"], "#,##0", fill=fill, align="right", bold=True, color="1668B8")
            cel(ws4, ri, 6,  g.get("fc",0), "#,##0", fill=fill, align="right")
            cel(ws4, ri, 7,  att,  "0.0%", fill=att_fill, align="center", bold=True,
                color="1AA15A" if on_t else ("D18A17" if att >= 0.5 else "D13B3B"))
            cel(ws4, ri, 8,  "✓" if on_t else "✗", fill=att_fill, align="center",
                bold=True, color="1AA15A" if on_t else "D13B3B")
            cel(ws4, ri, 9,  g.get("ig",0), "#,##0", fill=fill, align="right")
            cel(ws4, ri, 10, g.get("rg",0), "#,##0", fill=fill, align="right")
            cel(ws4, ri, 11, t.get("best_model", g.get("best_model","—")),
                fill=fill, align="center")
            cel(ws4, ri, 12, wm, "0.000", fill=fill, align="center",
                color="1AA15A" if wm < 0.12 else ("D18A17" if wm < 0.20 else "D13B3B"))
            cel(ws4, ri, 13, t.get("confidence", round(1-wm,2)), "0.0%", fill=fill, align="center")
            cel(ws4, ri, 14, ms, "0.00%", fill=fill, align="center")

        nr4 = len(best_r["goals"]) + 8
        ws4.merge_cells(start_row=nr4, start_column=1, end_row=nr4, end_column=4)
        c = ws4.cell(row=nr4, column=1, value="NATIONAL TOTAL / AVERAGE")
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill = HDR; c.alignment = Alignment(horizontal="center"); c.border = BDR
        for ci, (col, fmt, agg) in enumerate([(5,"#,##0","SUM"),(6,"#,##0","SUM"),
                                               (7,"0.0%","AVERAGE"),(9,"#,##0","SUM"),
                                               (10,"#,##0","SUM")], 0):
            actual_ci = [5,6,7,9,10][ci]
            cl = get_column_letter(actual_ci)
            ws4.cell(row=nr4, column=actual_ci, value=f"={agg}({cl}8:{cl}{nr4-1})")
            ws4.cell(row=nr4, column=actual_ci).font = Font(name="Arial", bold=True, color="FFFFFF")
            ws4.cell(row=nr4, column=actual_ci).fill = HDR
            ws4.cell(row=nr4, column=actual_ci).number_format = fmt
            ws4.cell(row=nr4, column=actual_ci).border = BDR
            ws4.cell(row=nr4, column=actual_ci).alignment = Alignment(horizontal="right")
        for col in [8,11,12,13,14]:
            ws4.cell(row=nr4, column=col).fill = HDR
            ws4.cell(row=nr4, column=col).border = BDR

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ── CSV export ────────────────────────────────────────────────────────────────
def to_csv(results: List[Dict]) -> bytes:
    out = io.StringIO()
    fields = [
        "scenario_id", "rank", "method", "mode",
        "rs", "is_", "ns", "w1", "w2", "w3",
        "composite", "confidence", "P", "A", "C", "Ga", "V", "K", "S",
        "territory_id", "territory", "region_id", "region",
        "final_goal", "forecast_fi", "attainment_pct",
        "is_best", "is_pref",
    ]
    writer = csv.DictWriter(out, fieldnames=fields)
    writer.writeheader()
    for r in results:
        sc, fs = r["scenario"], r["fairness"]
        for g in r["goals"]:
            writer.writerow({
                "scenario_id": g["sid"], "rank": r["rank"],
                "method": sc["method"], "mode": sc["mode"],
                "rs": sc["rs"], "is_": sc["is_"], "ns": sc["ns"],
                "w1": sc["w1"], "w2": sc["w2"], "w3": sc["w3"],
                "composite": fs["boosted_score"],
                "confidence": fs.get("confidence", ""),
                "P": fs["P"], "A": fs["A"], "C": fs["C"],
                "Ga": fs["Ga"], "V": fs["V"], "K": fs["K"], "S": fs["S"],
                "territory_id": g["tid"], "territory": g["tname"],
                "region_id": g["rid"], "region": g["rname"],
                "final_goal": g["fg"],
                "forecast_fi": g["fc"],
                "attainment_pct": round(g["att"] * 100, 1),
                "is_best": "YES" if r["is_best"] else "",
                "is_pref": "YES" if r.get("is_pref") else "",
            })
    return out.getvalue().encode("utf-8")
